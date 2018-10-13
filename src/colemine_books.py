#!/usr/bin/python

import datetime
import sys
import openpyxl
from config_stuff import *
from openpyxl.styles import NamedStyle, Font, Border, Side
import locale
import re
import collections


class ColemineBooks():
    def __init__(self, yr, q):
        self.wb = openpyxl.load_workbook(COLEMINE_BOOKS_FILE)
        self.data_input = self.wb.get_sheet_by_name(name = 'Data Input')
        self.projects = self.wb.get_sheet_by_name(name = 'Project & Royalties')
        locale.setlocale (locale.LC_ALL, '' )
        self.non_decimal = re.compile(r'[^\d.]+')
        self.year = yr
        self.quarter = q
        
        
    def get_projects(self):
        #get all projects
        p = self.projects['C']

        projects_set = set()
        for index, row in enumerate(p):
            if index == 0:
                continue
            if row.value:
                try:
                    projects_set.add(row.value.lower())
                    #print row.value.lower()
                except Exception as e:
                    projects_set.add(row.value)
                    #print row.value
                #print e
        sorted_projects = sorted(projects_set)
        #for row in sorted_projects:
        #    print row
        return sorted_projects

    def get_royalty_codes(self):
        #get all royalty codes
        r = self.projects['D']

        royalty_set = set()
        for row in r:
            if row.value:
                try:
                    royalty_set.add(row.value.lower())
                except Exception as e:
                    royalty_set.add(row.value)
                    #print e
        sorted_royalty = sorted(royalty_set)
        return sorted_royalty
        
        #print data_input
        #sheet_ranges = data_input['range names']
        #print sheet_ranges

    def get_payees(self):
        #p = self.projects['A']
        #codes
        payees_set = set()

        for index, row in enumerate(self.projects):
            if index == 0:
                continue
            if len(str(row[3].value)) < 2 or row[3].value is None:
                #this is a project summary, not a payee
                #print 'This is a summary, not a payee: %s - %s' % (str(row[0].value),str(row[1].value))
                continue
            try:
                payees_set.add(row[0].value.lower().strip())
            except Exception as e:
                print 'couldnt add payee to list: %s - %s' % (str(row[0].value), e)
        sorted_payees = sorted(payees_set)
        #print '--------------------------'
        #for row in sorted_payees:
            #print row
        #print '--------------------------'
        return sorted_payees
        
    def get_sales_categories(self):
        #get all categories
        categories = self.data_input['F']

        sales_categories = set()
        for row in categories:
            if row.value:
                try:
                    sales_categories.add(row.value.lower())
                except Exception as e:
                    print 'tried to add a none type to categories: %s' % e    
        #print sales_categories
        return sales_categories

    def get_format_categories(self):
        categories = self.data_input['E']

        format_categories = set()
        for row in categories:
            try:
                format_categories.add(str(row.value).lower())
            except Exception as e:
                print 'tried to add some fucked up format category: %s' % e
        return format_categories

    def get_project_status(self):
        project_codes = self.get_projects()
        royalty_codes = self.get_royalty_codes()
        project_status = {} #new dict
        for pro in project_codes:
            project_status[str(pro)] = 0.0
        #print project_status
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            #print '--%s--' % this_adjustment_project
            found = False
            if str(this_adjustment_project) in project_status:
                #print
                #print row[1]
                #print this_adjustment_project
                #print '---------'
                project_status[str(this_adjustment_project)] += float(eval(str(row[1].value).replace('=','')))
            else:
                if str(this_adjustment_project) not in royalty_codes:
                    print 'couldnt find matching key for row: %s %s --%s--' % (row[0].value, row[1].value, row[2].value)
                continue
        #for key, value in iter(sorted(project_status.iteritems())):
        #    print '%s - %s' % (key, str(value))
        return project_status

    def generate_label_royalty_lines(self):
        royalty_lines = []
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if str(this_adjustment_project)[0] == 'r':
                #print 'found some royalties'
                #print code
                #print '1'
                #print row
                #save summary line
                date = row[0].value
                amount = 0
                try:
                    amount = float(eval(str(row[1].value).replace('=','')))
                except Exception as e:
                    print 'warning: error in generate_label_royalty_lines'
                    print e
                    continue
                royalty = this_adjustment_project
                qty = row[3].value
                fmt = ''
                #print '2'
                try:
                    fmt = str(row[4].value).lower()
                except Exception as e:
                    placeholder = 0
                category = str(row[5].value).lower()
                #print '3'
                company = ''
                try:
                    company = self.xstr(row[6].value)
                except Exception as e:
                    placeholder = 0
                    #continue
                    #print '**********\n' * 5
                    #print row[5].value
                #print '4'
                details = str(row[7].value)
                details_2 = str(row[8].value)
                #print '5'
                #if this was a royalty line, get the summary and get the fuck outta here
                royalty_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                royalty_lines.append(royalty_line)
                #print '6'
                #print
        return royalty_lines
    
    
    def generate_royalty_lines(self, code):
        royalty_lines = []
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if (str(code) == str(this_adjustment_project)):
                #print 'found some royalties'
                #print code
                #print '1'
                #print row
                #save summary line
                date = row[0].value
                amount = float(eval(str(row[1].value).replace('=','')))
                royalty = this_adjustment_project
                qty = row[3].value
                fmt = ''
                #print '2'
                try:
                    fmt = str(row[4].value).lower()
                except Exception as e:
                    placeholder = 0
                category = str(row[5].value).lower()
                #print '3'
                company = ''
                try:
                    company = self.xstr(row[6].value)
                except Exception as e:
                    placeholder = 0
                    #continue
                    #print '**********\n' * 5
                    #print row[5].value
                #print '4'
                details = str(row[7].value)
                details_2 = str(row[8].value)
                #print '5'
                #if this was a royalty line, get the summary and get the fuck outta here
                royalty_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                royalty_lines.append(royalty_line)
                #print '6'
                #print
        return royalty_lines

    def generate_all_project_summary(self):
        dict_summary = {}
        dict_qty = {}
        dict_dict_summary = {}
        dict_dict_qty = {}
        dict_dict_title_summary = {}
        dict_dict_title_qty = {}
        summary_lines = []
        royalty_lines = []
        for item in self.get_sales_categories():
            dict_summary[str(item)] = 0.0
            dict_qty[str(item)] = 0
            dict_dict_summary[str(item)] = {}
            dict_dict_qty[str(item)] = {}
            for format_type in self.get_format_categories():
                dict_dict_summary[str(item)][str(format_type)] = 0.0
                dict_dict_qty[str(item)][str(format_type)] = 0
        for item in self.get_projects():
            dict_dict_title_summary[str(item)] = {}
            dict_dict_title_qty[str(item)] = {}
            #print item
            for format_type in self.get_format_categories():
                #print format_type
                dict_dict_title_summary[str(item)][str(format_type)] = 0.0
                dict_dict_title_qty[str(item)][str(format_type)] = 0
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            date = row[0].value
            amount = 0
            try:
                amount = float(eval(str(row[1].value).replace('=','')))
            except Exception as e:
                print 'trouble casting data input to amount:'
                print row[1].value
                print '---------'
                continue
            royalty = this_adjustment_project
            qty = row[3].value
            fmt = ''
            try:
                fmt = str(row[4].value).lower()
            except Exception as e:
                placeholder = 0
            category = str(row[5].value).lower()
            company = ''
            try:
                company = self.xstr(row[6].value)
            except Exception as e:
                placeholder = 0
                #continue
                #print '**********\n' * 5
                #print row[5].value
            details = str(row[7].value)
            details_2 = str(row[8].value)
            #if this was a royalty line, get the summary and get the fuck outta here
            if str(this_adjustment_project)[0] == 'r':
                royalty_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                royalty_lines.append(royalty_line)
                continue
            summary_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
            summary_lines.append(summary_line)
            #adjust dict by amount
            dict_summary[category] += amount
            if category in dict_dict_summary:
                if fmt in dict_dict_summary[category]:
                    dict_dict_summary[category][fmt] += amount
            #adjust qty if applicable
            if qty is not None:
                try:
                    qty = int(qty)
                    dict_qty[category] += qty
                    if category in dict_dict_summary:
                        if fmt in dict_dict_summary[category]:
                            dict_dict_qty[category][fmt] += qty
                except Exception as e:
                    print e
                    continue
            #for summary on each title
            if str(this_adjustment_project) in dict_dict_title_summary:
                if fmt in dict_dict_title_summary[str(this_adjustment_project)]:
                    dict_dict_title_summary[str(this_adjustment_project)][fmt] += amount
            #add info for qty sold by title and format
            if qty is not None:
                try:
                    qty = int(qty)
                    if str(this_adjustment_project) in dict_dict_title_qty:
                        if fmt in dict_dict_title_qty[str(this_adjustment_project)]:
                            dict_dict_title_qty[str(this_adjustment_project)][fmt] += qty
                except Exception as e:
                    print e
                    continue
        #for title in dict_dict_title_summary:
        #    for fmt, value in dict_dict_title_summary[title].iteritems():
        #        placeholder = 0
        #        print '%s\t%s\t%s' % (str(title), str(fmt), str(value))
        return (dict_summary, dict_qty, summary_lines, dict_dict_summary, dict_dict_qty, royalty_lines, dict_dict_title_summary, dict_dict_title_qty)


            
    def generate_project_summary(self, code):
        dict_summary = {}
        dict_qty = {}
        dict_dict_summary = {}
        dict_dict_qty = {}
        summary_lines = []
        royalty_lines = []
        for item in self.get_sales_categories():
            dict_summary[str(item)] = 0.0
            dict_qty[str(item)] = 0
            dict_dict_summary[str(item)] = {}
            dict_dict_qty[str(item)] = {}
            for format_type in self.get_format_categories():
                dict_dict_summary[str(item)][str(format_type)] = 0.0
                dict_dict_qty[str(item)][str(format_type)] = 0
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if code == str(this_adjustment_project) or (('r%s'%code) == str(this_adjustment_project)):
                #print row
                #save summary line
                date = row[0].value
                amount = float(eval(str(row[1].value).replace('=','')))
                royalty = this_adjustment_project
                qty = row[3].value
                fmt = ''
                try:
                    fmt = str(row[4].value).lower()
                except Exception as e:
                    placeholder = 0
                category = str(row[5].value).lower()
                company = ''
                try:
                    company = self.xstr(row[6].value)
                except Exception as e:
                    placeholder = 0
                    #continue
                    #print '**********\n' * 5
                    #print row[5].value
                details = str(row[7].value)
                details_2 = str(row[8].value)
                #if this was a royalty line, get the summary and get the fuck outta here
                if (('r%s'%code) == str(this_adjustment_project)):
                    royalty_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                    royalty_lines.append(royalty_line)
                    continue
                summary_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                summary_lines.append(summary_line)
                #adjust dict by amount
                dict_summary[category] += amount
                if category in dict_dict_summary:
                    if fmt in dict_dict_summary[category]:
                        dict_dict_summary[category][fmt] += amount
                #adjust qty if applicable
                if qty is not None:
                    try:
                        qty = int(qty)
                        dict_qty[category] += qty
                        if category in dict_dict_summary:
                            if fmt in dict_dict_summary[category]:
                                dict_dict_qty[category][fmt] += qty
                    except Exception as e:
                        continue
        return (dict_summary, dict_qty, summary_lines, dict_dict_summary, dict_dict_qty, royalty_lines)

    def generate_quarterly_label_project_summary(self, year, quarter):
        dict_summary = {}
        dict_qty = {}
        dict_dict_summary = {}
        dict_dict_qty = {}
        dict_dict_title_summary = {}
        dict_dict_title_qty = {}
        summary_lines = []
        for item in self.get_sales_categories():
            dict_summary[str(item)] = 0.0
            dict_qty[str(item)] = 0
            dict_dict_summary[str(item)] = {}
            dict_dict_qty[str(item)] = {}
            for format_type in self.get_format_categories():
                dict_dict_summary[str(item)][str(format_type)] = 0.0
                dict_dict_qty[str(item)][str(format_type)] = 0.0
        for item in self.get_projects():
            dict_dict_title_summary[str(item)] = {}
            dict_dict_title_qty[str(item)] = {}
            #print item
            for format_type in self.get_format_categories():
                #print format_type
                dict_dict_title_summary[str(item)][str(format_type)] = 0.0
                dict_dict_title_qty[str(item)][str(format_type)] = 0
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if str(this_adjustment_project)[0] is not 'r':
                #if code == str(this_adjustment_project):# or (('r%s'%code) == str(this_adjustment_project)):
                #print row
                #save summary line
                date = row[0].value
                #date_ = openpyxl.utils.datetime.from_excel(date)
                #date_ = datetime.datetime(date)
                #print date.month
                #print self.quarter
                #print
                #print date.year
                #print self.year
                #print
                if (((date.month-1)//3)+1) != int(quarter):
                    continue
                if date.year != int(year):
                    continue
                amount = float(eval(str(row[1].value).replace('=','')))
                royalty = this_adjustment_project
                qty = str(row[3].value)
                fmt = ''
                try:
                    fmt = str(row[4].value).lower()
                except Exception as e:
                    placeholder = 0
                category = str(row[5].value).lower()
                company = ''
                details = ''
                details_2 = ''
                try:
                    company = self.xstr(row[6].value)
                    details = str(row[7].value)
                    details_2 = str(row[8].value)
                except Exception as e:
                    placeholder = 0
                summary_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                summary_lines.append(summary_line)
                #adjust dict by amount
                dict_summary[category] += amount
                if category in dict_dict_summary:
                    if fmt in dict_dict_summary[category]:
                        dict_dict_summary[category][fmt] += amount
                #adjust qty if applicable
                if qty is not None:
                    try:
                        qty = int(qty)
                        dict_qty[category] += qty
                        if category in dict_dict_summary:
                            if fmt in dict_dict_summary[category]:
                                dict_dict_qty[category][fmt] += qty
                    except Exception as e:
                        continue
                #for summary on each title
                if str(this_adjustment_project) in dict_dict_title_summary:
                    if fmt in dict_dict_title_summary[str(this_adjustment_project)]:
                        dict_dict_title_summary[str(this_adjustment_project)][fmt] += amount
                #add info for qty sold by title and format
                if qty is not None:
                    try:
                        qty = int(qty)
                        if str(this_adjustment_project) in dict_dict_title_qty:
                            if fmt in dict_dict_title_qty[str(this_adjustment_project)]:
                                dict_dict_title_qty[str(this_adjustment_project)][fmt] += qty
                    except Exception as e:
                        print e
                        continue
                    
        return (dict_summary, dict_qty, summary_lines, dict_dict_summary, dict_dict_qty, dict_dict_title_summary, dict_dict_title_qty)
    
    
    def generate_quarterly_project_summary(self, code, year, quarter):
        dict_summary = {}
        dict_qty = {}
        dict_dict_summary = {}
        dict_dict_qty = {}
        summary_lines = []
        for item in self.get_sales_categories():
            dict_summary[str(item)] = 0.0
            dict_qty[str(item)] = 0
            dict_dict_summary[str(item)] = {}
            dict_dict_qty[str(item)] = {}
            for format_type in self.get_format_categories():
                dict_dict_summary[str(item)][str(format_type)] = 0.0
                dict_dict_qty[str(item)][str(format_type)] = 0.0
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if code == str(this_adjustment_project):# or (('r%s'%code) == str(this_adjustment_project)):
                #print row
                #save summary line
                date = row[0].value
                #date_ = openpyxl.utils.datetime.from_excel(date)
                #date_ = datetime.datetime(date)
                #print date.month
                #print self.quarter
                #print
                #print date.year
                #print self.year
                #print
                if (((date.month-1)//3)+1) != int(quarter):
                    continue
                if date.year != int(year):
                    continue
                amount = float(eval(str(row[1].value).replace('=','')))
                royalty = this_adjustment_project
                qty = str(row[3].value)
                fmt = ''
                try:
                    fmt = str(row[4].value).lower()
                except Exception as e:
                    placeholder = 0
                category = str(row[5].value).lower()
                company = ''
                details = ''
                details_2 = ''
                try:
                    company = self.xstr(row[6].value)
                    details = str(row[7].value)
                    details_2 = str(row[8].value)
                except Exception as e:
                    placeholder = 0
                summary_line = [date, amount, royalty, qty, fmt, category, company, details, details_2]
                summary_lines.append(summary_line)
                #adjust dict by amount
                dict_summary[category] += amount
                if category in dict_dict_summary:
                    if fmt in dict_dict_summary[category]:
                        dict_dict_summary[category][fmt] += amount
                #adjust qty if applicable
                if qty is not None:
                    try:
                        qty = int(qty)
                        dict_qty[category] += qty
                        if category in dict_dict_summary:
                            if fmt in dict_dict_summary[category]:
                                dict_dict_qty[category][fmt] += qty
                    except Exception as e:
                        continue
        return (dict_summary, dict_qty, summary_lines, dict_dict_summary, dict_dict_qty)

    def generate_label_royalty_summary(self):
        royalty_summary = []
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            if str(this_adjustment_project)[0] == 'r':
                royalty_type = ''
                product_qty = 0
                product_format = ''
                amount = 0
                try:
                    amount = float(eval(str(row[1].value).replace('=','')))
                except Exception as e:
                    print 'warning: error in generate_label_royalty_summary'
                    print e
                    continue
                if row[5].value == 'Royalties (Product)':
                    royalty_type = 'Product'
                    product_qty = 0
                    try:
                        product_qty = int(row[3].value)
                    except Exception as e:
                        placeholder = 0
                    product_format = str(row[4].value)
                else:
                    royalty_type = 'Cash'
                royalty_summary.append([royalty_type, amount, product_format, product_qty])
        return royalty_summary


    def generate_royalty_summary(self, code):
        #print '--------'
        #print code
        royalty_summary = []
        for row in self.data_input:
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            found = False
            if this_adjustment_project == code:
                royalty_type = ''
                product_qty = 0
                product_format = ''
                amount = float(eval(str(row[1].value).replace('=','')))
                if row[5].value == 'Royalties (Product)':
                    royalty_type = 'Product'
                    product_qty = 0
                    try:
                        product_qty = int(row[3].value)
                    except Exception as e:
                        placeholder = 0
                    product_format = str(row[4].value)
                else:
                    royalty_type = 'Cash'
                royalty_summary.append([royalty_type, amount, product_format, product_qty])
        return royalty_summary
    
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thick', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border
    
    def get_royalty_status(self):
        royalty_codes = self.get_royalty_codes()
        project_codes = self.get_projects()
        project_codes = map(str, project_codes)
        royalty_status = {}
        for roy in royalty_codes:
            royalty_status[str(roy)] = 0.0
        for row in self.data_input:
            #print row
            this_adjustment_project = row[2].value
            try:
                this_adjustment_project = this_adjustment_project.lower()
            except Exception as e:
                placeholder = 0
            found = False
            if str(this_adjustment_project) in royalty_status:
                royalty_status[str(this_adjustment_project)] += float(eval(str(row[1].value).replace('=','')))
            else:
                if str(this_adjustment_project) not in project_codes:
                    print 'couldnt find matching key for row: %s %s %s' % (row[0].value, row[1].value, row[2].value)
                    #print project_codes
                continue
        #for key, value in iter(sorted(royalty_status.iteritems())):
        #    print '%s - %s' % (key, str(value))
        return royalty_status

            #adjustments = self.data_input['B']
            #for index, item in adjustments:
    def correct_column_widths(self, wb):
        for sheet in wb.worksheets:
            for col in ['A','B','C','D','E','F','G','H','I','J']:
                sheet_col = sheet[col]
                list_of_lengths = []
                list_of_lengths.append(1)
                for row in sheet_col[1:50]:
                    try:
                        list_of_lengths.append(len(str(row.value)))
                    except Exception as e:
                        placeholder = 0
                max_length = max(list_of_lengths)+2
                max_length = min(max_length, 50)
                sheet.column_dimensions[col].width = max_length


    def print_label_summary(self):
        project_status = self.get_project_status()
        royalty_status = self.get_royalty_status()
        project_sheet_row = 1
        wb = openpyxl.Workbook()
        ws_project = wb.active
        ws_project.title = "Label Summary"
        summary_of_payments = []

        #giant project workbook
        project_summary = self.generate_all_project_summary()

        (ws_project.cell(row=project_sheet_row, column=1)).value = 'LABEL MASTER'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 1
        (ws_project.cell(row=project_sheet_row, column=1)).value = 'ALL PROJECTS'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 1
        (ws_project.cell(row=project_sheet_row, column=1)).value = 'CAT NO NOT APPLICABLE'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 1
        (ws_project.cell(row=project_sheet_row, column=1)).value = 'WE DON\'T MAKE MONEY'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 2

        need_details_for = set()
        need_sub_details_for = set()

        
        #royalty summaries -------------------------------
        project_sheet_details_row = 1
        royalty_summary = self.generate_label_royalty_summary()
        royalty_lines = self.generate_label_royalty_lines()
        #print royalty_lines
        types_of_royalties = set()
        types_amount = {}
        types_qty = {}
        for royalty_row in royalty_summary:
            if royalty_row[0] == 'Cash':
                types_of_royalties.add(royalty_row[0])
            else:
                types_of_royalties.add('%s - %s' % (royalty_row[0], royalty_row[2]))
        for types in types_of_royalties:
            types_amount[types] = 0.0
            types_qty[types] = 0
        for royalty_row in royalty_summary:
            if royalty_row[0] == 'Cash':
                types_amount[royalty_row[0]] += royalty_row[1]
                types_qty[royalty_row[0]] += royalty_row[3]
            else:#product
                type_of_product = '%s - %s' % (royalty_row[0], royalty_row[2])
                types_amount[type_of_product] += royalty_row[1]
                types_qty[type_of_product] += royalty_row[3]
        #print royalty_number
        if len(types_of_royalties) > 0:
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'ROYALTIES PAID'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            self.set_border(ws_project, 'E1:E1')
            project_sheet_details_row += 1
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'Royalty Type'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_details_row, column=6)).value = 'Amount'
            (ws_project.cell(row=project_sheet_details_row, column=6)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_details_row, column=7)).value = 'QTY'
            (ws_project.cell(row=project_sheet_details_row, column=7)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1
            amount_sums = 0
            qty_sums = 0
            for index, types, in enumerate(types_of_royalties):
                (ws_project.cell(row=project_sheet_details_row, column=5)).value = types
                (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_details_row, column=6)).value = -types_amount[types]
                (ws_project.cell(row=project_sheet_details_row, column=6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                (ws_project.cell(row=project_sheet_details_row, column=7)).value = types_qty[types]
                amount_sums += -types_amount[types]
                qty_sums += types_qty[types]
                project_sheet_details_row += 1
            (ws_project.cell(row=project_sheet_details_row, column=6)).value = amount_sums
            (ws_project.cell(row=project_sheet_details_row, column=6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            (ws_project.cell(row=project_sheet_details_row, column=7)).value = qty_sums
            self.set_border(ws_project, 'E2:G%s' % int(project_sheet_details_row))
            project_sheet_details_row += 3

        #fix rows so that quarters match up with each other and shit #fuckformatting
        project_sheet_row = max(project_sheet_row, project_sheet_details_row)
        project_sheet_details_row = max(project_sheet_row, project_sheet_details_row)


        #end royalty summaries ---------------

        start_row = project_sheet_row
        (ws_project.cell(row=project_sheet_row, column=1)).value = 'ALL TIME SUMMARY'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 1
        for index, col, in enumerate(('Category', 'Debit/Credit', 'QTY')):
            (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
            (ws_project.cell(row=project_sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
        project_sheet_row += 1

                    
        total_sold = 0
        project_budget = 0
        budget_positive = 0
        budget_negative = 0
        qty_positive = 0
        qty_negative = 0
        pos_summary = False
        #for key, value in sorted(project_summary[0].iteritems()):
        for key, value in sorted(project_summary[0].items(), key=lambda x: x[1],reverse=True):
            if float(value) == 0:#we've reached the center of the list, do things
                if budget_positive > 0 and not pos_summary:#display summary
                    (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Income'
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    (ws_project.cell(row=project_sheet_row, column=2)).value = budget_positive
                    (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                    (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    (ws_project.cell(row=project_sheet_row, column=3)).value = qty_positive
                    (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 2
                    pos_summary = True
                    continue
            project_budget += value
            total_sold += (project_summary[1])[key]
            if value > 0:
                budget_positive += value
                qty_positive += (project_summary[1])[key]
            elif value < 0:
                budget_negative += value
                qty_negative += (project_summary[1])[key]
            (ws_project.cell(row=project_sheet_row, column=1)).value = str(key).title()
            (ws_project.cell(row=project_sheet_row, column=2)).value = value
            (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            (ws_project.cell(row=project_sheet_row, column=3)).value = (project_summary[1])[key]
            project_sheet_row += 1

            #display format sub-categories if necessary
            if key in project_summary[3]: #does this category have sub-catgeories
                #print '%s - yep, shes got sub categories' % catalog_number
                #next line loops through the sub-categories of a category in sorted order
                for key_sub, value_sub in sorted((project_summary[3])[key].items(), key=lambda x: x[1], reverse=True):
                    #print '%s - %s - %s - %s' % (catalog_number, key, key_sub, str(value_sub))
                    if float(value_sub) == 0:
                        continue
                    if key_sub == 'none' or key_sub is None or key_sub == '':
                        continue
                    need_details_for.add(key)
                    need_sub_details_for.add(key_sub)
                    category = '   - %s %s' % (key, key_sub)
                    #(ws_project.cell(row=project_sheet_row, column=1)).value = str(category).title()
                    #(ws_project.cell(row=project_sheet_row, column=2)).value = value_sub
                    #(ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    #(ws_project.cell(row=project_sheet_row, column=3)).value = (project_summary[4])[key][key_sub]
                    #project_sheet_row += 1
                        
        #display negative summaries
        if budget_negative < 0:#display summary
            (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Expenses'
            (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_row, column=2)).value = budget_negative
            (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            #(ws_project.cell(row=project_sheet_row, column=3)).value = qty_negative
            #(ws_project.cell(row=project_sheet_row, column=4)).font = openpyxl.styles.Font(bold=True)
            project_sheet_row += 2                    
            #display neg and pos combined    
        (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Income Minus Expenses'
        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
        (ws_project.cell(row=project_sheet_row, column=2)).value = project_budget
        (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
        (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
        (ws_project.cell(row=project_sheet_row, column=3)).value = total_sold
        (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)

        #create border
        self.set_border(ws_project, ('A%s:A%s' % (str(start_row), str(start_row))))
        self.set_border(ws_project, ('A%s:C%s' % (str(start_row+1), str(project_sheet_row))))

        label_summary_details_row = project_sheet_details_row
        
       #display detailed breakdown of sales with sub-categories -------------
        need_sub_details_for = list(need_sub_details_for)
        need_sub_details_for.sort()
        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'ALL TIME BY FORMAT'
        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
        self.set_border(ws_project, 'E%s:E%s' % (str(project_sheet_details_row), str(project_sheet_details_row)))
        project_sheet_details_row += 1
        start_row = project_sheet_details_row
        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'INCOME BREAKDOWN'
        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
        sub_sums = {}
        for index, col, in enumerate(need_sub_details_for):
            sub_sums[col] = 0.0
            if len(str(col)) == 2:
                col = str(col).upper()
            else:
                col = str(col).title()
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
            (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
        project_sheet_details_row += 1
        for key in list(need_details_for):
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            category_total = 0.0
            end_column = 20
            for index, key_sub, in enumerate(need_sub_details_for):
                category_total += project_summary[3][key][key_sub]
                sub_sums[key_sub] += project_summary[3][key][key_sub]
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_summary[3][key][key_sub]
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                end_column = index+7
            (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = category_total
            (ws_project.cell(row=project_sheet_details_row, column=end_column)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1
        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
        for index, col, in enumerate(need_sub_details_for):
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = sub_sums[col]
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
        project_sheet_details_row += 1

                            

        project_sheet_details_row += 1
        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'QTY BREAKDOWN'
        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
        qty_sums = {}
        for index, col, in enumerate(need_sub_details_for):
            qty_sums[col] = 0.0
            if len(str(col)) == 2:
                col = str(col).upper()
            else:
                col = str(col).title()
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
            (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
        project_sheet_details_row += 1
        column_for_border = 0
        for key in list(need_details_for):
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            qty_total = 0.0
            end_column = 20
            for index, key_sub, in enumerate(need_sub_details_for):
                qty_sums[key_sub] += project_summary[4][key][key_sub]
                qty_total += project_summary[4][key][key_sub]
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_summary[4][key][key_sub]
                #(ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                end_column = index+7
            (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = qty_total
            (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1
            column_for_border = end_column - 1
        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
        self.set_border(ws_project, ('E%s:%s%s' % (str(start_row), chr(ord('A') + column_for_border),str(project_sheet_details_row))))
        for index, col, in enumerate(need_sub_details_for):
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = qty_sums[col]
            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
        project_sheet_details_row += 1

        #start of title breakdown ---------------------
        (ws_project.cell(row=label_summary_details_row, column=14)).value = 'ALL TIME TITLE BREAKDOWN SUMMARY'
        (ws_project.cell(row=label_summary_details_row, column=14)).font = openpyxl.styles.Font(bold=True)
        self.set_border(ws_project, 'N%s:Q%s' % (str(label_summary_details_row),str(label_summary_details_row)))
        self.set_border(ws_project, 'N%s:Q%s' % (str(label_summary_details_row+1),str(label_summary_details_row+1)))
        start_label_summary_details_row = label_summary_details_row
        label_summary_details_row += 1

        for index, col in enumerate(('Title', 'Format', 'Amount', 'Qty')):
            (ws_project.cell(row=label_summary_details_row, column=index+14)).value = col
            (ws_project.cell(row=label_summary_details_row, column=index+14)).font = openpyxl.styles.Font(bold=True)
        label_summary_details_row += 1

        label_display_array = []
        for title in project_summary[6]:
            for fmt, value in project_summary[6][title].iteritems():
                if fmt in ('cd', 'cs', 'lp', '45'):
                    if value > 0:
                        label_display_array.append([str(title),str(fmt),str(value),project_summary[7][title][fmt]])
        label_display_array = sorted(label_display_array, key = lambda x: x[3], reverse=True)
        print label_display_array

        for row in label_display_array:
            (ws_project.cell(row=label_summary_details_row, column=14)).value = row[0]
            (ws_project.cell(row=label_summary_details_row, column=15)).value = row[1]
            (ws_project.cell(row=label_summary_details_row, column=16)).value = locale.currency(float(row[2]))
            (ws_project.cell(row=label_summary_details_row, column=16)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            (ws_project.cell(row=label_summary_details_row, column=17)).value = row[3]
            label_summary_details_row += 1
                #(ws_project.cell(row=label_summary_details_row, column=index+14)).font = openpyxl.styles.Font(bold=True)
        self.set_border(ws_project, 'N%s:Q%s' % (str(start_label_summary_details_row),str(label_summary_details_row)))
        label_summary_details_row = start_label_summary_details_row
        #end of title breakdown ------------------

        
        #quarterly summary ------------------
        num_quarters = 4
        pairs = []
        year = int(self.year)
        quarter = int(self.quarter)
        for ii in range(num_quarters):
            if quarter == 0:
                quarter = 4
                year -= 1
            pairs.append([year,quarter])
            quarter -= 1
        for pair_index, pair in enumerate(pairs):
            #fix rows so that quarters match up with each other and shit #fuckformatting
            project_sheet_row = max(project_sheet_row, project_sheet_details_row)
            project_sheet_details_row = max(project_sheet_row, project_sheet_details_row)
                    
            project_quarterly_summary = self.generate_quarterly_label_project_summary(int(pair[0]),int(pair[1]))
                        
            need_details_for = set()
            need_sub_details_for = set()
                        
            project_sheet_row += 2
            border_start_quarter = project_sheet_row
            (ws_project.cell(row=project_sheet_row, column=1)).value = '%s Q%s SUMMARY' % (str(pair[0]), str(pair[1]))
            (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
            project_sheet_row += 1
            for index, col, in enumerate(('Category', 'Debit/Credit', 'QTY')):
                (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                (ws_project.cell(row=project_sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
            project_sheet_row += 1
            total_sold = 0
            project_budget = 0
            #for key, value in sorted(project_summary[0].iteritems()):
            budget_positive = 0
            budget_negative = 0
            qty_positive = 0
            qty_negative = 0
            pos_summary = False
            for key, value in sorted(project_quarterly_summary[0].items(), key=lambda x: x[1],reverse=True):
                if float(value) == 0:#we've reached the center of the list, do things
                    if budget_positive > 0 and not pos_summary:#display summary
                        (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Income'
                        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_row, column=2)).value = budget_positive
                        (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        (ws_project.cell(row=project_sheet_row, column=3)).value = qty_positive
                        (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_row += 2
                        pos_summary = True
                    continue
                project_budget += value
                total_sold += (project_quarterly_summary[1])[key]
                if value > 0:
                    budget_positive += value
                    qty_positive += (project_quarterly_summary[1])[key]
                elif value < 0:
                    budget_negative += value
                    qty_negative += (project_quarterly_summary[1])[key]
                (ws_project.cell(row=project_sheet_row, column=1)).value = str(key).title()
                (ws_project.cell(row=project_sheet_row, column=2)).value = value
                (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                (ws_project.cell(row=project_sheet_row, column=3)).value = (project_quarterly_summary[1])[key]
                project_sheet_row += 1

                if key in project_quarterly_summary[3]: #does this category have sub-catgeories
                    #print '%s - yep, shes got sub categories' % catalog_number
                    #next line loops through the sub-categories of a category in sorted order
                    for key_sub, value_sub in sorted((project_quarterly_summary[3])[key].items(), key=lambda x: x[1], reverse=True):
                        if float(value_sub) == 0:
                            continue
                        if key_sub == 'none' or key_sub is None or key_sub == '':
                            continue
                        need_details_for.add(key)
                        need_sub_details_for.add(key_sub)
                                    
            #display negative summaries
            if budget_negative < 0:#display summary
                (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Expenses'
                (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_row, column=2)).value = budget_negative
                (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                #(ws_project.cell(row=project_sheet_row, column=3)).value = qty_negative
                #(ws_project.cell(row=project_sheet_row, column=4)).font = openpyxl.styles.Font(bold=True)
                project_sheet_row += 2                    

            #display positive minus negative
            (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Totals'
            (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_row, column=2)).value = project_budget
            (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
            (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            (ws_project.cell(row=project_sheet_row, column=3)).value = total_sold
            (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)

            #create border
            self.set_border(ws_project, ('A%s:C%s' % (str(border_start_quarter),str(project_sheet_row))))

            #quarterly summary detailed breakdown ---------------

            #display detailed breakdown of sales with sub-categories -------------
            need_sub_details_for = list(need_sub_details_for)
            need_sub_details_for.sort()
            #print need_sub_details_for
            project_sheet_details_row += 2
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = '%s Q%s BY FORMAT' % (str(pair[0]), str(pair[1]))
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            self.set_border(ws_project, 'E%s:E%s' % (int(project_sheet_details_row),int(project_sheet_details_row)))
            project_sheet_details_row += 1
            start_quarterly_details_border = project_sheet_details_row #save this row for later
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'INCOME BREAKDOWN'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            sub_sums = {}
            for index, col, in enumerate(need_sub_details_for):
                sub_sums[col] = 0.0
                if len(str(col)) == 2:
                    col = str(col).upper()
                else:
                    col = str(col).title()
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1
            for key in list(need_details_for):
                (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                category_total = 0.0
                end_column = 20
                for index, key_sub, in enumerate(need_sub_details_for):
                    category_total += project_quarterly_summary[3][key][key_sub]
                    sub_sums[key_sub] += project_quarterly_summary[3][key][key_sub]
                    (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_quarterly_summary[3][key][key_sub]
                    (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    end_column = index+7
                (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = category_total
                (ws_project.cell(row=project_sheet_details_row, column=end_column)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                project_sheet_details_row += 1
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
            for index, col, in enumerate(need_sub_details_for):
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = sub_sums[col]
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            project_sheet_details_row += 1

                            

            project_sheet_details_row += 1
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'QTY BREAKDOWN'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
            qty_sums = {}
            for index, col, in enumerate(need_sub_details_for):
                qty_sums[col] = 0.0
                if len(str(col)) == 2:
                    col = str(col).upper()
                else:
                    col = str(col).title()
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1
            column_for_border = 0
            for key in list(need_details_for):
                (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                qty_total = 0.0
                end_column = 20
                for index, key_sub, in enumerate(need_sub_details_for):
                    qty_sums[key_sub] += project_quarterly_summary[4][key][key_sub]
                    qty_total += project_quarterly_summary[4][key][key_sub]
                    (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_quarterly_summary[4][key][key_sub]
                    end_column = index+7
                (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = qty_total
                (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                project_sheet_details_row += 1
                column_for_border = end_column - 1
            (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
            self.set_border(ws_project, ('E%s:%s%s' % (start_quarterly_details_border,chr(ord('A') + column_for_border),str(project_sheet_details_row))))
            for index, col, in enumerate(need_sub_details_for):
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = qty_sums[col]
                (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
            project_sheet_details_row += 1

            #quarterly summary detailed breakdown DONE ----------------



            #start of title breakdown ---------------------
            alpha = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z','aa','ab','ac','ad','ae','af','ag','ah','ai','aj','ak','al','am','an','ao','ap','aq','ar','as','at','au','av','aw','ax','ay','az','ba','bb','bc','bd','be','bf','bg','bh','bi','bj','bk','bl','bm','bn','bo','bp','bq','br','bs','bt','bu','bv','bw','bx','by','bz','ca','cb','cc','cd','ce','cf','cg','ch','ci','cj','ck','cl','cm','cn','co','cp','cq','cr','cs','ct','cu','cv','cw','cx','cy','cz']

            (ws_project.cell(row=label_summary_details_row, column=19+(6*pair_index))).value = '%s Q%s BY TITLE' % (str(pair[0]), str(pair[1]))
            (ws_project.cell(row=label_summary_details_row, column=19+(6*pair_index))).font = openpyxl.styles.Font(bold=True)
            self.set_border(ws_project, '%s%s:%s%s' % (str(alpha[18+6*pair_index]),str(label_summary_details_row),str(alpha[18+6*pair_index+4]),str(label_summary_details_row)))
            self.set_border(ws_project, '%s%s:%s%s' % (str(alpha[18+6*pair_index]),str(label_summary_details_row+1),str(alpha[18+6*pair_index+4]),str(label_summary_details_row+1)))
            start_label_summary_details_row = label_summary_details_row
            label_summary_details_row += 1

            for index, col in enumerate(('Title', 'Format', 'Amount', 'Qty', '$ Per Unit')):
                (ws_project.cell(row=label_summary_details_row, column=19+6*pair_index+index)).value = col
                (ws_project.cell(row=label_summary_details_row, column=19+6*pair_index+index)).font = openpyxl.styles.Font(bold=True)
            label_summary_details_row += 1

            label_display_array = []
            for title in project_quarterly_summary[5]:
                for fmt, value in project_quarterly_summary[5][title].iteritems():
                    if fmt in ('cd', 'cs', 'lp', '45'):
                        if value > 0:
                            label_display_array.append([str(title),str(fmt),str(value),project_quarterly_summary[6][title][fmt]])
            label_display_array = sorted(label_display_array, key = lambda x: x[3], reverse=True)
            print label_display_array
        
            for row in label_display_array:
                (ws_project.cell(row=label_summary_details_row, column=19+6*pair_index)).value = row[0]
                (ws_project.cell(row=label_summary_details_row, column=20+6*pair_index)).value = row[1]
                (ws_project.cell(row=label_summary_details_row, column=21+6*pair_index)).value = locale.currency(float(row[2]))
                (ws_project.cell(row=label_summary_details_row, column=21+6*pair_index)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                (ws_project.cell(row=label_summary_details_row, column=22+6*pair_index)).value = row[3]
                (ws_project.cell(row=label_summary_details_row, column=23+6*pair_index)).value = locale.currency(float(row[2])/float(row[3]))
                (ws_project.cell(row=label_summary_details_row, column=23+6*pair_index)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                label_summary_details_row += 1

            self.set_border(ws_project, '%s%s:%s%s' % (alpha[18+6*pair_index],str(start_label_summary_details_row),alpha[18+6*pair_index+4],str(label_summary_details_row)))
            label_summary_details_row = start_label_summary_details_row
        
        self.correct_column_widths(wb)
        wb.save(BASE_PATH + 'plaid_room/royalty_summaries/LABEL_MASTER.xlsx')

            
    def print_payee_summary(self):
        project_status = self.get_project_status()
        royalty_status = self.get_royalty_status()
        payees = self.get_payees()
        sheet_row_master = 1
        wb_master = openpyxl.Workbook()
        ws_master = wb_master.active
        ws_master.title = "Master Summary"
        summary_of_payments = []
        for payee in payees:
            sheet_row = 1
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "All Projects Summary"
            summary_of_payments = []
            #print '-----------------------------------------------'
            #print payee
            for row in self.projects:
                if str(row[0].value).lower().strip() == payee:#this guy is on this project, what we owe him?
                    #first make sure that this is an actual payee, and not a project summary
                    if len(str(row[3].value)) < 2 or row[3].value is None:
                        #this is a project summary, not a payee
                        continue
                    #get project number and validate
                    catalog_number = str(row[2].value)
                    try:
                        catalog_number = catalog_number.lower()
                    except Exception as e:
                        placeholder = 0
                    if catalog_number not in project_status:
                        print 'Holy shit, this dude is on a project that doesn\'t even exist'
                        print catalog_number
                        print row[0].value
                        print row
                        continue
                    #what is this payee's percentage?
                    percent = float(row[5].value)
                    project_budget_status = float(project_status[catalog_number])
                    money_owed =  project_budget_status * percent
                    #get royalty code and validate
                    money_paid = 0
                    royalty_number = row[3].value
                    try:
                        royalty_number = royalty_number.lower()
                    except Exception as e:
                        placeholder = 0
                    if royalty_number not in royalty_status:
                        print 'Invalid royalty number or this person has never been paid - %s' % royalty_number
                    else:
                        money_paid = float(royalty_status[royalty_number])
                    money_owed_after_paid = money_owed + money_paid
                    summary_of_payments.append([str(payee),str(row[1].value),str(row[2].value),str(row[3].value),project_budget_status,str('%.2f'%(percent*100))+'%',money_owed,-money_paid,money_owed_after_paid])
                    
                    #individual project workbooks
                    #generate project_summary
                    project_summary = self.generate_project_summary(catalog_number)

                    need_details_for = set()
                    need_sub_details_for = set()
                    sheet_name = str(row[1].value).replace(' ','_')
                    sheet_name = re.sub(r'\W+', '', sheet_name)
                    ws_project = wb.create_sheet('%s_%s' % (str(row[2].value), sheet_name))
                    project_sheet_row = 1
                    #create header
                    #finding the artist is hard, here goes
                    p = self.projects['C']
                    artist = ''
                    for index, row_ in enumerate(p):
                        if str(row_.value) == str(catalog_number):
                            artist = self.projects['A%s' % str(index+1)].value
                            break
                    (ws_project.cell(row=project_sheet_row, column=1)).value = artist.title()
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1
                    (ws_project.cell(row=project_sheet_row, column=1)).value = str(row[1].value)
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1
                    catalog_no = catalog_number
                    if str(catalog_number).isdigit():
                        catalog_no = 'CLMN-%s' % str(catalog_number)
                    (ws_project.cell(row=project_sheet_row, column=1)).value = str(catalog_no)
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1
                    (ws_project.cell(row=project_sheet_row, column=1)).value = 'Artist Profit Split - ' + str('%.2f'%(percent*100)) + '%'
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 2
                    #done making header
                    #royalty summaries -------------------------------
                    project_sheet_details_row = 1
                    royalty_summary = self.generate_royalty_summary(royalty_number)#[type amount format qty]
                    royalty_lines = self.generate_royalty_lines(royalty_number)
                    #print royalty_lines
                    types_of_royalties = set()
                    types_amount = {}
                    types_qty = {}
                    for royalty_row in royalty_summary:
                        if royalty_row[0] == 'Cash':
                            types_of_royalties.add(royalty_row[0])
                        else:
                            types_of_royalties.add('%s - %s' % (royalty_row[0], royalty_row[2]))
                    for types in types_of_royalties:
                        types_amount[types] = 0.0
                        types_qty[types] = 0
                    for royalty_row in royalty_summary:
                        if royalty_row[0] == 'Cash':
                            types_amount[royalty_row[0]] += royalty_row[1]
                            types_qty[royalty_row[0]] += royalty_row[3]
                        else:#product
                            type_of_product = '%s - %s' % (royalty_row[0], royalty_row[2])
                            types_amount[type_of_product] += royalty_row[1]
                            types_qty[type_of_product] += royalty_row[3]
                    #print royalty_number
                    #for types in types_of_royalties:
                    #    print '%s - %s - %s' % (types, -types_amount[types], types_qty[types])
                    #print
                    if len(types_of_royalties) > 0:
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'ROYALTIES PAID'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        self.set_border(ws_project, 'E1:E1')
                        project_sheet_details_row += 1
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'Royalty Type'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_details_row, column=6)).value = 'Amount'
                        (ws_project.cell(row=project_sheet_details_row, column=6)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_details_row, column=7)).value = 'QTY'
                        (ws_project.cell(row=project_sheet_details_row, column=7)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1
                        amount_sums = 0
                        qty_sums = 0
                        for index, types, in enumerate(types_of_royalties):
                            (ws_project.cell(row=project_sheet_details_row, column=5)).value = types
                            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_details_row, column=6)).value = -types_amount[types]
                            (ws_project.cell(row=project_sheet_details_row, column=6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            (ws_project.cell(row=project_sheet_details_row, column=7)).value = types_qty[types]
                            amount_sums += -types_amount[types]
                            qty_sums += types_qty[types]
                            project_sheet_details_row += 1
                        (ws_project.cell(row=project_sheet_details_row, column=6)).value = amount_sums
                        (ws_project.cell(row=project_sheet_details_row, column=6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        (ws_project.cell(row=project_sheet_details_row, column=7)).value = qty_sums
                        self.set_border(ws_project, 'E2:G%s' % int(project_sheet_details_row))
                        project_sheet_details_row += 3

                    #fix rows so that quarters match up with each other and shit #fuckformatting
                    project_sheet_row = max(project_sheet_row, project_sheet_details_row)
                    project_sheet_details_row = max(project_sheet_row, project_sheet_details_row)

                    #print royalty_number
                    #print royalty_summary

                    
                    #end royalty summaries ---------------

                    start_row = project_sheet_row
                    (ws_project.cell(row=project_sheet_row, column=1)).value = 'ALL TIME SUMMARY'
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1
                    for index, col, in enumerate(('Category', 'Debit/Credit', 'QTY')):
                        (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                        (ws_project.cell(row=project_sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1

                    
                    total_sold = 0
                    project_budget = 0
                    budget_positive = 0
                    budget_negative = 0
                    qty_positive = 0
                    qty_negative = 0
                    pos_summary = False
                    #for key, value in sorted(project_summary[0].iteritems()):
                    for key, value in sorted(project_summary[0].items(), key=lambda x: x[1],reverse=True):
                        if float(value) == 0:#we've reached the center of the list, do things
                            if budget_positive > 0 and not pos_summary:#display summary
                                (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Income'
                                (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                                (ws_project.cell(row=project_sheet_row, column=2)).value = budget_positive
                                (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                                (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                                (ws_project.cell(row=project_sheet_row, column=3)).value = qty_positive
                                (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)
                                project_sheet_row += 2
                                pos_summary = True
                            continue
                        project_budget += value
                        total_sold += (project_summary[1])[key]
                        if value > 0:
                            budget_positive += value
                            qty_positive += (project_summary[1])[key]
                        elif value < 0:
                            budget_negative += value
                            qty_negative += (project_summary[1])[key]
                        (ws_project.cell(row=project_sheet_row, column=1)).value = str(key).title()
                        (ws_project.cell(row=project_sheet_row, column=2)).value = value
                        (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        (ws_project.cell(row=project_sheet_row, column=3)).value = (project_summary[1])[key]
                        project_sheet_row += 1

                        #display format sub-categories if necessary
                        if key in project_summary[3]: #does this category have sub-catgeories
                            #print '%s - yep, shes got sub categories' % catalog_number
                            #next line loops through the sub-categories of a category in sorted order
                            for key_sub, value_sub in sorted((project_summary[3])[key].items(), key=lambda x: x[1], reverse=True):
                                #print '%s - %s - %s - %s' % (catalog_number, key, key_sub, str(value_sub))
                                if float(value_sub) == 0:
                                    continue
                                if key_sub == 'none' or key_sub is None or key_sub == '':
                                    continue
                                need_details_for.add(key)
                                need_sub_details_for.add(key_sub)
                                category = '   - %s %s' % (key, key_sub)
                                #(ws_project.cell(row=project_sheet_row, column=1)).value = str(category).title()
                                #(ws_project.cell(row=project_sheet_row, column=2)).value = value_sub
                                #(ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                                #(ws_project.cell(row=project_sheet_row, column=3)).value = (project_summary[4])[key][key_sub]
                                #project_sheet_row += 1
                        
                    #display negative summaries
                    if budget_negative < 0:#display summary
                                (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Expenses'
                                (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                                (ws_project.cell(row=project_sheet_row, column=2)).value = budget_negative
                                (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                                (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                                #(ws_project.cell(row=project_sheet_row, column=3)).value = qty_negative
                                #(ws_project.cell(row=project_sheet_row, column=4)).font = openpyxl.styles.Font(bold=True)
                                project_sheet_row += 2                    
                    #display neg and pos combined    
                    (ws_project.cell(row=project_sheet_row, column=1)).value = 'Total Income Minus Expenses'
                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                    (ws_project.cell(row=project_sheet_row, column=2)).value = project_budget
                    (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                    (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    (ws_project.cell(row=project_sheet_row, column=3)).value = total_sold
                    (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)

                    #create border
                    self.set_border(ws_project, ('A%s:A%s' % (str(start_row), str(start_row))))
                    self.set_border(ws_project, ('A%s:C%s' % (str(start_row+1), str(project_sheet_row))))

                    #display detailed breakdown of sales with sub-categories -------------
                    need_sub_details_for = list(need_sub_details_for)
                    need_sub_details_for.sort()
                    (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'ALL TIME BY FORMAT'
                    (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                    self.set_border(ws_project, 'E%s:E%s' % (str(project_sheet_details_row), str(project_sheet_details_row)))
                    project_sheet_details_row += 1
                    start_row = project_sheet_details_row
                    (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'INCOME BREAKDOWN'
                    (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                    sub_sums = {}
                    for index, col, in enumerate(need_sub_details_for):
                        sub_sums[col] = 0.0
                        if len(str(col)) == 2:
                            col = str(col).upper()
                        else:
                            col = str(col).title()
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                        (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_details_row += 1
                    for key in list(need_details_for):
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        category_total = 0.0
                        end_column = 20
                        for index, key_sub, in enumerate(need_sub_details_for):
                            category_total += project_summary[3][key][key_sub]
                            sub_sums[key_sub] += project_summary[3][key][key_sub]
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_summary[3][key][key_sub]
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            end_column = index+7
                        (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = category_total
                        (ws_project.cell(row=project_sheet_details_row, column=end_column)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1
                    (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
                    (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
                    for index, col, in enumerate(need_sub_details_for):
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = sub_sums[col]
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                    project_sheet_details_row += 1

                            

                    project_sheet_details_row += 1
                    (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'QTY BREAKDOWN'
                    (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                    qty_sums = {}
                    for index, col, in enumerate(need_sub_details_for):
                        qty_sums[col] = 0.0
                        if len(str(col)) == 2:
                            col = str(col).upper()
                        else:
                            col = str(col).title()
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                        (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_details_row += 1
                    column_for_border = 0
                    for key in list(need_details_for):
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        qty_total = 0.0
                        end_column = 20
                        for index, key_sub, in enumerate(need_sub_details_for):
                            qty_sums[key_sub] += project_summary[4][key][key_sub]
                            qty_total += project_summary[4][key][key_sub]
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_summary[4][key][key_sub]
                            #(ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            end_column = index+7
                        (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = qty_total
                        (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1
                        column_for_border = end_column - 1
                    (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
                    (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
                    self.set_border(ws_project, ('E%s:%s%s' % (str(start_row), chr(ord('A') + column_for_border),str(project_sheet_details_row))))
                    for index, col, in enumerate(need_sub_details_for):
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = qty_sums[col]
                        (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_details_row += 1

                    #quarterly summary ---------------
                    num_quarters = 4
                    pairs = []
                    year = int(self.year)
                    quarter = int(self.quarter)
                    for ii in range(num_quarters):
                        if quarter == 0:
                            quarter = 4
                            year -= 1
                        pairs.append([year,quarter])
                        quarter -= 1
                    for pair in pairs:
                        #fix rows so that quarters match up with each other and shit #fuckformatting
                        project_sheet_row = max(project_sheet_row, project_sheet_details_row)
                        project_sheet_details_row = max(project_sheet_row, project_sheet_details_row)
                        
                        project_quarterly_summary = self.generate_quarterly_project_summary(catalog_number,int(pair[0]),int(pair[1]))
                        
                        need_details_for = set()
                        need_sub_details_for = set()
                        
                        project_sheet_row += 2
                        border_start_quarter = project_sheet_row
                        (ws_project.cell(row=project_sheet_row, column=1)).value = '%s Q%s SUMMARY' % (str(pair[0]), str(pair[1]))
                        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_row += 1
                        for index, col, in enumerate(('Category', 'Debit/Credit', 'QTY')):
                            (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                            (ws_project.cell(row=project_sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_row += 1
                        total_sold = 0
                        project_budget = 0
                        #for key, value in sorted(project_summary[0].iteritems()):
                        budget_positive = 0
                        budget_negative = 0
                        qty_positive = 0
                        qty_negative = 0
                        pos_summary = False
                        for key, value in sorted(project_quarterly_summary[0].items(), key=lambda x: x[1],reverse=True):
                            if float(value) == 0:#we've reached the center of the list, do things
                                if budget_positive > 0 and not pos_summary:#display summary
                                    (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Income'
                                    (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                                    (ws_project.cell(row=project_sheet_row, column=2)).value = budget_positive
                                    (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                                    (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                                    (ws_project.cell(row=project_sheet_row, column=3)).value = qty_positive
                                    (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)
                                    project_sheet_row += 2
                                    pos_summary = True
                                continue
                            project_budget += value
                            total_sold += (project_quarterly_summary[1])[key]
                            if value > 0:
                                budget_positive += value
                                qty_positive += (project_quarterly_summary[1])[key]
                            elif value < 0:
                                budget_negative += value
                                qty_negative += (project_quarterly_summary[1])[key]
                            (ws_project.cell(row=project_sheet_row, column=1)).value = str(key).title()
                            (ws_project.cell(row=project_sheet_row, column=2)).value = value
                            (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            (ws_project.cell(row=project_sheet_row, column=3)).value = (project_quarterly_summary[1])[key]
                            project_sheet_row += 1

                            if key in project_quarterly_summary[3]: #does this category have sub-catgeories
                                #print '%s - yep, shes got sub categories' % catalog_number
                                #next line loops through the sub-categories of a category in sorted order
                                for key_sub, value_sub in sorted((project_quarterly_summary[3])[key].items(), key=lambda x: x[1], reverse=True):
                                    if float(value_sub) == 0:
                                        continue
                                    if key_sub == 'none' or key_sub is None or key_sub == '':
                                        continue
                                    need_details_for.add(key)
                                    need_sub_details_for.add(key_sub)
                            
                        #display negative summaries
                        if budget_negative < 0:#display summary
                            (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Expenses'
                            (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_row, column=2)).value = budget_negative
                            (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            #(ws_project.cell(row=project_sheet_row, column=3)).value = qty_negative
                            #(ws_project.cell(row=project_sheet_row, column=4)).font = openpyxl.styles.Font(bold=True)
                            project_sheet_row += 2                    

                        #display positive minus negative
                        (ws_project.cell(row=project_sheet_row, column=1)).value = 'Quarterly Totals'
                        (ws_project.cell(row=project_sheet_row, column=1)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_row, column=2)).value = project_budget
                        (ws_project.cell(row=project_sheet_row, column=2)).font = openpyxl.styles.Font(bold=True)
                        (ws_project.cell(row=project_sheet_row, column=2)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        (ws_project.cell(row=project_sheet_row, column=3)).value = total_sold
                        (ws_project.cell(row=project_sheet_row, column=3)).font = openpyxl.styles.Font(bold=True)

                        #create border
                        self.set_border(ws_project, ('A%s:C%s' % (str(border_start_quarter),str(project_sheet_row))))

                        #quarterly summary detailed breakdown ---------------

                        #display detailed breakdown of sales with sub-categories -------------
                        need_sub_details_for = list(need_sub_details_for)
                        need_sub_details_for.sort()
                        #print need_sub_details_for
                        project_sheet_details_row += 2
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = '%s Q%s BY FORMAT' % (str(pair[0]), str(pair[1]))
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        self.set_border(ws_project, 'E%s:E%s' % (int(project_sheet_details_row),int(project_sheet_details_row)))
                        project_sheet_details_row += 1
                        start_quarterly_details_border = project_sheet_details_row #save this row for later
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'INCOME BREAKDOWN'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        sub_sums = {}
                        for index, col, in enumerate(need_sub_details_for):
                            sub_sums[col] = 0.0
                            if len(str(col)) == 2:
                                col = str(col).upper()
                            else:
                                col = str(col).title()
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                            (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1
                        for key in list(need_details_for):
                            (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                            category_total = 0.0
                            end_column = 20
                            for index, key_sub, in enumerate(need_sub_details_for):
                                category_total += project_quarterly_summary[3][key][key_sub]
                                sub_sums[key_sub] += project_quarterly_summary[3][key][key_sub]
                                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_quarterly_summary[3][key][key_sub]
                                (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                                end_column = index+7
                            (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = category_total
                            (ws_project.cell(row=project_sheet_details_row, column=end_column)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                            project_sheet_details_row += 1
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
                        for index, col, in enumerate(need_sub_details_for):
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = sub_sums[col]
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        project_sheet_details_row += 1

                            

                        project_sheet_details_row += 1
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'QTY BREAKDOWN'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                        qty_sums = {}
                        for index, col, in enumerate(need_sub_details_for):
                            qty_sums[col] = 0.0
                            if len(str(col)) == 2:
                                col = str(col).upper()
                            else:
                                col = str(col).title()
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = col
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                            (ws_project.cell(row=project_sheet_details_row, column=index+7)).value = 'TOTAL'
                            (ws_project.cell(row=project_sheet_details_row, column=index+7)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1
                        column_for_border = 0
                        for key in list(need_details_for):
                            (ws_project.cell(row=project_sheet_details_row, column=5)).value = str(key).title()
                            (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)
                            qty_total = 0.0
                            end_column = 20
                            for index, key_sub, in enumerate(need_sub_details_for):
                                qty_sums[key_sub] += project_quarterly_summary[4][key][key_sub]
                                qty_total += project_quarterly_summary[4][key][key_sub]
                                (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = project_quarterly_summary[4][key][key_sub]
                                end_column = index+7
                            (ws_project.cell(row=project_sheet_details_row, column=end_column)).value = qty_total
                            (ws_project.cell(row=project_sheet_details_row, column=end_column)).font = openpyxl.styles.Font(bold=True)
                            project_sheet_details_row += 1
                            column_for_border = end_column - 1
                        (ws_project.cell(row=project_sheet_details_row, column=5)).value = 'TOTAL'
                        (ws_project.cell(row=project_sheet_details_row, column=5)).font = openpyxl.styles.Font(bold=True)       
                        self.set_border(ws_project, ('E%s:%s%s' % (start_quarterly_details_border,chr(ord('A') + column_for_border),str(project_sheet_details_row))))
                        for index, col, in enumerate(need_sub_details_for):
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).value = qty_sums[col]
                            (ws_project.cell(row=project_sheet_details_row, column=index+6)).font = openpyxl.styles.Font(bold=True)
                        project_sheet_details_row += 1

                        #quarterly summary detailed breakdown DONE ----------------
                        

                    
                    
                    #all the data ------------
                    project_sheet_row += 3
                    project_sheet_details_row += 3

                    #fix rows so that quarters match up with each other and shit #fuckformatting
                    project_sheet_row = max(project_sheet_row, project_sheet_details_row)
                    project_sheet_details_row = max(project_sheet_row, project_sheet_details_row)
                    for index, col, in enumerate(('Date', 'Credit/Debit', 'Royalty', 'Qty','Format', 'Category','Company','Details','Details 2')):
                        (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                        (ws_project.cell(row=project_sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
                    project_sheet_row += 1
                    for row_ in reversed(project_summary[2]):
                        for index, col in enumerate(row_):
                            (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                            if index == 0:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).number_format = 'm/d/yyyy;@'
                                #(ws_project.cell(row=project_sheet_row, column=index+1)).number_format = '[$-409]m/d/yy h:mm AM/PM;@'
                            if index == 1:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            if index == 5:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).value = col.title()
                        project_sheet_row += 1
                    #print royalty_lines
                    #print '---------------'
                    for row_ in reversed(royalty_lines):
                        for index, col in enumerate(row_):
                            (ws_project.cell(row=project_sheet_row, column=index+1)).value = col
                            if index == 0:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).number_format = 'm/d/yyyy;@'
                                #(ws_project.cell(row=project_sheet_row, column=index+1)).number_format = '[$-409]m/d/yy h:mm AM/PM;@'
                            if index == 1:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                            if index == 5:
                                (ws_project.cell(row=project_sheet_row, column=index+1)).value = col.title()
                        project_sheet_row += 1
                    
            #print '|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|' % ('Payee','Title','Catalog No', 'Code', 'Total', 'Percentage','Total Earned','Total Paid','Current Due')
            for index, col, in enumerate(('Payee','Title','Catalog No', 'Code', 'Project Status', 'Percentage','Total Earned','Total Paid','Current Due')):
                (ws.cell(row=sheet_row, column=index+1)).value = col
                (ws.cell(row=sheet_row, column=index+1)).font = openpyxl.styles.Font(bold=True)
            sheet_row += 1

            for index, col, in enumerate(('Payee','Title','Catalog No', 'Code', 'Project Status', 'Percentage','Total Earned','Total Paid','Current Due')):
                (ws_master.cell(row=sheet_row_master, column=index+1)).value = col
            sheet_row_master += 1
            
            total_owed = 0
            for row in summary_of_payments:
                #print '|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|' % tuple(row) 
                #print '--------------------------------------------------------------------------' * 2
                for index, col in enumerate(row):
                    if isinstance(col, str):
                        (ws.cell(row=sheet_row, column=index+1)).value = col.title()
                    else:
                        (ws.cell(row=sheet_row, column=index+1)).value = col                        
                    if (index+1) in [5, 7, 8, 9]:#rows with $$ formatting needed
                        (ws.cell(row=sheet_row, column=index+1)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                        if (index+1) in [7,9]:
                            if col < 0:
                                (ws.cell(row=sheet_row, column=index+1)).value = 0                   
                                
                sheet_row += 1
                #cash_me_outside = self.xfloat(row[-1].replace('$',''))
                cash_me_outside = row[-1]
                if cash_me_outside > 0:
                    total_owed += cash_me_outside
            (ws.cell(row=sheet_row,column=8)).value = 'Sum Of Projects With Royalties Owed'
            #(ws.cell(row=sheet_row,column=9)).style = 'Currency' 
            (ws.cell(row=sheet_row,column=9)).value = str(locale.currency(total_owed))
            (ws.cell(row=sheet_row,column=9)).font = openpyxl.styles.Font(bold=True)
            sheet_row += 2
            #print '-------------------------------------'
            self.correct_column_widths(wb)
            wb.save(BASE_PATH + 'plaid_room/royalty_summaries/%s.xlsx' % (payee.replace(' ','-')))

            #master summary workbook
            total_owed = 0
            for row in summary_of_payments:
                #print '|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|\t%s\t|' % tuple(row) 
                for index, col in enumerate(row):
                    (ws_master.cell(row=sheet_row_master, column=index+1)).value = col
                    if (index+1) in [5, 7, 8, 9]:#rows with $$ formatting needed
                        (ws_master.cell(row=sheet_row_master, column=index+1)).number_format = '$#,##0.00;[Red]-$#,##0.00'
                sheet_row_master += 1
                #cash_me_outside = self.xfloat(row[-1].replace('$',''))
                cash_me_outside = row[-1]
                if cash_me_outside > 0:
                    total_owed += cash_me_outside
            (ws_master.cell(row=sheet_row_master,column=8)).value = 'Sum Of Projects With Royalties Owed'
            (ws_master.cell(row=sheet_row_master,column=8)).font = openpyxl.styles.Font(bold=True)
            #(ws.cell(row=sheet_row,column=9)).style = 'Currency' 
            #(ws_master.cell(row=sheet_row_master,column=9)).value = str(locale.currency(total_owed))
            (ws_master.cell(row=sheet_row_master,column=9)).value = total_owed
            (ws_master.cell(row=sheet_row_master,column=9)).font = openpyxl.styles.Font(bold=True)
            (ws_master.cell(row=sheet_row_master,column=9)).number_format = '$#,##0.00;[Red]-$#,##0.00'
            sheet_row_master += 2
            self.correct_column_widths(wb_master)
            
        wb_master.save(BASE_PATH + 'plaid_room/royalty_summaries/MASTER_FOR_COLEMINE_EMPLOYEES.xlsx')
            
                    
                    
    def filter_non_numeric(self, str_):
        return self.non_decimal.sub('', str_)

    def xfloat(self, f):
        if (f is None) or (f == ''):
            return -1
        return float(f)

    def filter_unprintable(self, str_):
        return filter(lambda x: x in string.printable, str_)

    
    def print_sheet_names(self):
        for sheet in self.wb.worksheets:
            print sheet

    def xstr(self,s):
        if s is None:
            return ''
        return str(s.encode('utf-8'))



if __name__ == '__main__':
    books = ColemineBooks(sys.argv[1], sys.argv[2])
    #books.print_sheet_names()
    #books.get_sales_categories()
    #books.get_projects()
    #books.get_project_status()
    #books.get_payees()
    books.print_label_summary()
    print 'done with label summary'
    books.print_payee_summary()
